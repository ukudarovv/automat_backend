import csv
import urllib.parse
from datetime import timedelta
from typing import Optional

from django.contrib.auth import login, logout
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q, Count, Avg, F
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect, render, get_object_or_404
from django.utils import timezone
from django.views.decorators.http import require_http_methods
from django.contrib import messages

from accounts.models import SrmUser
from leads.models import Lead, LeadStatusHistory
from settings_app.models import WhatsAppTemplate, ProjectSetting
from .forms import LeadFilterForm, LeadStatusUpdateForm


def login_view(request):
    """Страница входа"""
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        from django.contrib.auth import authenticate
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            return redirect("srm_ui:dashboard")
        else:
            messages.error(request, "Неверное имя пользователя или пароль")
    return render(request, "srm_ui/login.html")


@login_required
def logout_view(request):
    """Выход"""
    logout(request)
    return redirect("srm_ui:login")


def _get_filtered_queryset(request):
    """Получить отфильтрованный queryset лидов с учетом прав доступа"""
    queryset = Lead.objects.all()
    user = request.user
    
    # SCHOOL_MANAGER видит только свои лиды
    if hasattr(user, "role") and user.role == SrmUser.Roles.SCHOOL_MANAGER:
        if user.school:
            queryset = queryset.filter(school=user.school)
        else:
            queryset = Lead.objects.none()
    
    # Применяем фильтры
    form = LeadFilterForm(request.GET)
    if form.is_valid():
        # Множественный выбор статусов
        if form.cleaned_data.get("status"):
            queryset = queryset.filter(status__in=form.cleaned_data["status"])
        if form.cleaned_data.get("type"):
            queryset = queryset.filter(type=form.cleaned_data["type"])
        if form.cleaned_data.get("city"):
            queryset = queryset.filter(city=form.cleaned_data["city"])
        if form.cleaned_data.get("category"):
            queryset = queryset.filter(category=form.cleaned_data["category"])
        if form.cleaned_data.get("school"):
            queryset = queryset.filter(school=form.cleaned_data["school"])
        if form.cleaned_data.get("instructor"):
            queryset = queryset.filter(instructor=form.cleaned_data["instructor"])
        if form.cleaned_data.get("language"):
            queryset = queryset.filter(language=form.cleaned_data["language"])
        if form.cleaned_data.get("source"):
            queryset = queryset.filter(source=form.cleaned_data["source"])
        if form.cleaned_data.get("search"):
            search_term = form.cleaned_data["search"]
            queryset = queryset.filter(
                Q(name__icontains=search_term) |
                Q(phone__icontains=search_term) |
                Q(iin__icontains=search_term) |
                Q(email__icontains=search_term) |
                Q(whatsapp__icontains=search_term)
            )
        if form.cleaned_data.get("created_from"):
            queryset = queryset.filter(created_at__gte=form.cleaned_data["created_from"])
        if form.cleaned_data.get("created_to"):
            queryset = queryset.filter(created_at__lte=form.cleaned_data["created_to"])
    
    # Сортировка по дате создания (новые сверху)
    queryset = queryset.order_by("-created_at")
    return queryset, form


@login_required
def lead_list(request):
    """Список лидов с фильтрацией и пагинацией"""
    # Экспорт Excel
    if request.GET.get("export") == "excel":
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            messages.error(request, "Библиотека openpyxl не установлена. Установите: pip install openpyxl")
            return redirect("srm_ui:lead_list")
        
        queryset, _ = _get_filtered_queryset(request)
        
        wb = openpyxl.Workbook()
        
        # Лист для всех лидов
        ws = wb.active
        ws.title = "Все лиды"
        
        # Заголовки
        headers = [
            "ID", "Тип", "Статус", "Имя", "Телефон", "ИИН", "WhatsApp", "Email",
            "Город", "Категория", "Автошкола", "Инструктор", "Тариф", "Цена", "Дата создания"
        ]
        
        # Стили для заголовков
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Данные
        for row_num, lead in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=str(lead.id))
            ws.cell(row=row_num, column=2, value=lead.get_type_display())
            ws.cell(row=row_num, column=3, value=lead.get_status_display())
            ws.cell(row=row_num, column=4, value=lead.name)
            ws.cell(row=row_num, column=5, value=lead.phone)
            ws.cell(row=row_num, column=6, value=lead.iin or "")
            ws.cell(row=row_num, column=7, value=lead.whatsapp or "")
            ws.cell(row=row_num, column=8, value=lead.email or "")
            ws.cell(row=row_num, column=9, value=lead.city.name_ru if lead.city else "")
            ws.cell(row=row_num, column=10, value=lead.category.name_ru if lead.category else "")
            ws.cell(row=row_num, column=11, value=lead.school.name_ru if lead.school else "")
            ws.cell(row=row_num, column=12, value=lead.instructor.display_name if lead.instructor else "")
            ws.cell(row=row_num, column=13, value=lead.tariff_plan.code if lead.tariff_plan else "")
            ws.cell(row=row_num, column=14, value=lead.tariff_price_kzt or lead.instructor_tariff_price_kzt or "")
            ws.cell(row=row_num, column=15, value=lead.created_at.strftime("%d.%m.%Y %H:%M"))
        
        # Автоподбор ширины колонок
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width
        
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="leads.xlsx"'
        wb.save(response)
        return response
    
    # Экспорт CSV
    if request.GET.get("export") == "csv":
        queryset, _ = _get_filtered_queryset(request)
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="leads.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            "ID", "Тип", "Статус", "Имя", "Телефон", "ИИН", "WhatsApp", "Email",
            "Город", "Категория", "Автошкола", "Инструктор", "Дата создания"
        ])
        
        for lead in queryset:
            writer.writerow([
                str(lead.id),
                lead.get_type_display(),
                lead.get_status_display(),
                lead.name,
                lead.phone,
                lead.iin or "",
                lead.whatsapp or "",
                lead.email or "",
                lead.city.name_ru if lead.city else "",
                lead.category.name_ru if lead.category else "",
                lead.school.name_ru if lead.school else "",
                lead.instructor.display_name if lead.instructor else "",
                lead.created_at.strftime("%d.%m.%Y %H:%M"),
            ])
        
        return response
    
    # Обычный список
    queryset, form = _get_filtered_queryset(request)
    
    # Пагинация
    paginator = Paginator(queryset, 20)
    page_number = request.GET.get("page", 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        "page_obj": page_obj,
        "form": form,
        "total_count": queryset.count(),
    }
    return render(request, "srm_ui/lead_list.html", context)


@login_required
def lead_detail(request, lead_id):
    """Детальный просмотр лида"""
    lead = get_object_or_404(Lead, id=lead_id)
    
    # Проверка прав доступа
    user = request.user
    if hasattr(user, "role") and user.role == SrmUser.Roles.SCHOOL_MANAGER:
        if not user.school or lead.school != user.school:
            messages.error(request, "У вас нет доступа к этому лиду")
            return redirect("srm_ui:lead_list")
    
    # История статусов
    status_history = lead.status_history.all()
    
    # Форма изменения статуса
    status_form = LeadStatusUpdateForm(initial={"new_status": lead.status})
    
    # WhatsApp шаблоны для выбора
    whatsapp_templates = []
    if lead.type == Lead.LeadType.SCHOOL:
        scope = WhatsAppTemplate.Scope.SCHOOL_CLIENT_MESSAGE
    elif lead.type == Lead.LeadType.INSTRUCTOR:
        scope = WhatsAppTemplate.Scope.INSTRUCTOR_CLIENT_MESSAGE
    elif lead.type == Lead.LeadType.TESTS:
        scope = WhatsAppTemplate.Scope.TESTS_OWNER_MESSAGE
    else:
        scope = None
    
    if scope:
        whatsapp_templates = WhatsAppTemplate.objects.filter(scope=scope, is_active=True)
    
    # Комментарии
    from .models import LeadComment
    comments = LeadComment.objects.filter(lead=lead).order_by("-created_at")
    
    context = {
        "lead": lead,
        "status_history": status_history,
        "status_form": status_form,
        "whatsapp_templates": whatsapp_templates,
        "comments": comments,
    }
    return render(request, "srm_ui/lead_detail.html", context)


@login_required
@require_http_methods(["POST"])
def lead_status_update(request, lead_id):
    """Обновление статуса лида"""
    lead = get_object_or_404(Lead, id=lead_id)
    
    # Проверка прав доступа
    user = request.user
    if hasattr(user, "role") and user.role == SrmUser.Roles.SCHOOL_MANAGER:
        if not user.school or lead.school != user.school:
            messages.error(request, "У вас нет доступа к этому лиду")
            return redirect("srm_ui:lead_list")
    
    form = LeadStatusUpdateForm(request.POST)
    if form.is_valid():
        old_status = lead.status
        new_status = form.cleaned_data["new_status"]
        note = form.cleaned_data.get("note", "")
        
        if old_status != new_status:
            lead.status = new_status
            lead.save()
            
            # Создаем запись в истории
            LeadStatusHistory.objects.create(
                lead=lead,
                old_status=old_status,
                new_status=new_status,
                changed_by_user=user,
                note=note,
            )
            
            # Отслеживаем ключевые события для аналитики
            if new_status == Lead.LeadStatus.PAID:
                # Отслеживаем оплату
                from analytics.models import AnalyticsEvent
                AnalyticsEvent.objects.create(
                    lead=lead,
                    bot_user=lead.bot_user,
                    event_name="lead_paid",
                    payload={"old_status": old_status, "new_status": new_status},
                )
            elif new_status == Lead.LeadStatus.DONE:
                # Отслеживаем завершение услуги
                from analytics.models import AnalyticsEvent
                AnalyticsEvent.objects.create(
                    lead=lead,
                    bot_user=lead.bot_user,
                    event_name="lead_completed",
                    payload={"old_status": old_status, "new_status": new_status},
                )
            
            messages.success(request, f"Статус изменен: {old_status} → {new_status}")
        else:
            messages.info(request, "Статус не изменился")
    
    return redirect("srm_ui:lead_detail", lead_id=lead_id)


def _build_wa_link(phone: str, text: str) -> str:
    """Построить WhatsApp ссылку"""
    if not phone:
        return ""
    clean_phone = phone.replace("+", "").replace(" ", "").replace("-", "")
    return f"https://wa.me/{clean_phone}?text={urllib.parse.quote(text)}"


def _get_whatsapp_template(lead: Lead) -> Optional[WhatsAppTemplate]:
    """Получить WhatsApp шаблон для лида"""
    scope = None
    if lead.type == Lead.LeadType.SCHOOL:
        scope = WhatsAppTemplate.Scope.SCHOOL_CLIENT_MESSAGE
    elif lead.type == Lead.LeadType.INSTRUCTOR:
        scope = WhatsAppTemplate.Scope.INSTRUCTOR_CLIENT_MESSAGE
    elif lead.type == Lead.LeadType.TESTS:
        scope = WhatsAppTemplate.Scope.TESTS_OWNER_MESSAGE
    
    if not scope:
        return None
    
    try:
        return WhatsAppTemplate.objects.get(scope=scope, language=lead.language, is_active=True)
    except WhatsAppTemplate.DoesNotExist:
        return None


def _generate_whatsapp_message(lead: Lead) -> str:
    """Генерировать WhatsApp сообщение для лида"""
    template = _get_whatsapp_template(lead)
    
    if template:
        # Используем шаблон из БД
        kwargs = {
            "name": lead.name,
            "phone": lead.phone,
        }
        
        if lead.type == Lead.LeadType.SCHOOL and lead.school:
            kwargs["school_name"] = lead.school.name_ru if lead.language == "RU" else lead.school.name_kz
            kwargs["address"] = lead.school.address_ru if lead.language == "RU" else lead.school.address_kz
            if lead.school.nearest_intake_date:
                kwargs["intake_date"] = lead.school.nearest_intake_date.strftime("%d.%m.%Y")
            elif lead.school.nearest_intake_text_ru:
                kwargs["intake_date"] = lead.school.nearest_intake_text_ru if lead.language == "RU" else lead.school.nearest_intake_text_kz
            else:
                kwargs["intake_date"] = "Ближайший набор" if lead.language == "RU" else "Келесі набор"
            
            if lead.tariff_plan and lead.tariff_price_kzt:
                tariff_names = {1: {"RU": "Базовый", "KZ": "Негізгі"}, 2: {"RU": "Стандарт", "KZ": "Стандарт"}, 3: {"RU": "Премиум", "KZ": "Премиум"}}
                tariff_name = tariff_names.get(lead.tariff_plan.id, {}).get(lead.language, lead.tariff_plan.code)
                kwargs["tariff"] = f"{tariff_name} {lead.tariff_price_kzt} KZT"
            else:
                kwargs["tariff"] = ""
        elif lead.type == Lead.LeadType.INSTRUCTOR and lead.instructor:
            kwargs["instructor_name"] = lead.instructor.display_name
            kwargs["gearbox"] = lead.gearbox or ""
            if lead.instructor_tariff and lead.instructor_tariff_price_kzt:
                kwargs["tariff"] = f"{lead.instructor_tariff.name_ru if lead.language == 'RU' else lead.instructor_tariff.name_kz} {lead.instructor_tariff_price_kzt} KZT"
            else:
                kwargs["tariff"] = f"{lead.instructor.price_kzt} KZT" if lead.instructor.price_kzt else ""
        elif lead.type == Lead.LeadType.TESTS:
            kwargs["iin"] = lead.iin or ""
            kwargs["whatsapp"] = lead.whatsapp or lead.phone
            if lead.category:
                kwargs["category"] = lead.category.name_ru if lead.language == "RU" else (lead.category.name_kz or lead.category.name_ru)
            else:
                kwargs["category"] = ""
        
        return template.render(**kwargs)
    else:
        # Fallback - генерируем сообщение вручную
        if lead.type == Lead.LeadType.SCHOOL and lead.school:
            school_name = lead.school.name_ru if lead.language == "RU" else lead.school.name_kz
            address = lead.school.address_ru if lead.language == "RU" else lead.school.address_kz
            intake = ""
            if lead.school.nearest_intake_date:
                intake = lead.school.nearest_intake_date.strftime("%d.%m.%Y")
            elif lead.school.nearest_intake_text_ru:
                intake = lead.school.nearest_intake_text_ru if lead.language == "RU" else lead.school.nearest_intake_text_kz
            
            tariff_text = ""
            if lead.tariff_plan and lead.tariff_price_kzt:
                tariff_names = {1: {"RU": "Базовый", "KZ": "Негізгі"}, 2: {"RU": "Стандарт", "KZ": "Стандарт"}, 3: {"RU": "Премиум", "KZ": "Премиум"}}
                tariff_name = tariff_names.get(lead.tariff_plan.id, {}).get(lead.language, lead.tariff_plan.code)
                tariff_text = f"{tariff_name} {lead.tariff_price_kzt} KZT"
            
            if lead.language == "KZ":
                return (
                    f"Құрметті болашақ жүргізуші!\n"
                    f"Сіз {school_name} таңдадыңыз\n"
                    f"📅 Топтың басталу күні: {intake or 'Келесі набор'}\n"
                    f"📍 Мекенжайы: {address}\n"
                    f"💰 Бағасы: {tariff_text}\n\n"
                    f"Қосымша сұрақтарыңыз бар ма\n"
                    f"әлде төлем жасауға дайынсыз ба?"
                )
            else:
                return (
                    f"Уважаемый(-ая), будущий водитель!\n"
                    f"Вы выбрали автошколу {school_name}\n"
                    f"📅 Начало группы: {intake or 'Ближайший набор'}\n"
                    f"📍 Адрес: {address}\n"
                    f"💰 Цена: {tariff_text}\n\n"
                    f"У вас есть дополнительные вопросы\n"
                    f"или вы готовы перейти к оплате?"
                )
        elif lead.type == Lead.LeadType.INSTRUCTOR and lead.instructor:
            instructor_name = lead.instructor.display_name
            price = lead.instructor.price_kzt
            if lead.language == "KZ":
                return f"Сәлем! Мен {lead.name}. Нұсқаушы {instructor_name} ({price} KZT) таңдадым."
            else:
                return f"Здравствуйте! Я {lead.name}. Выбрал инструктора {instructor_name} ({price} KZT)."
        else:
            return f"Здравствуйте! Меня зовут {lead.name}. Телефон: {lead.phone}"


@login_required
@require_http_methods(["POST"])
def lead_payment_link_update(request, lead_id):
    """Обновить ссылку на оплату"""
    lead = get_object_or_404(Lead, id=lead_id)
    
    # Проверка прав доступа
    user = request.user
    if hasattr(user, "role") and user.role == SrmUser.Roles.SCHOOL_MANAGER:
        if not user.school or lead.school != user.school:
            messages.error(request, "У вас нет доступа к этому лиду")
            return redirect("srm_ui:lead_list")
    
    payment_link = request.POST.get("payment_link", "").strip()
    if payment_link:
        lead.payment_link = payment_link
        lead.save()
        messages.success(request, "Ссылка на оплату сохранена")
    else:
        lead.payment_link = None
        lead.save()
        messages.success(request, "Ссылка на оплату удалена")
    
    return redirect("srm_ui:lead_detail", lead_id=lead_id)


@login_required
def dashboard(request):
    """Дашборд с аналитикой и метриками"""
    user = request.user
    queryset = Lead.objects.all()
    
    # Фильтрация по правам доступа
    if hasattr(user, "role") and user.role == SrmUser.Roles.SCHOOL_MANAGER:
        if user.school:
            queryset = queryset.filter(school=user.school)
        else:
            queryset = Lead.objects.none()
    
    # Базовые метрики
    total_leads = queryset.count()
    
    # Лиды по статусам
    leads_by_status = queryset.values('status').annotate(count=Count('id')).order_by('status')
    status_counts = {item['status']: item['count'] for item in leads_by_status}
    
    # Лиды по типам
    leads_by_type = queryset.values('type').annotate(count=Count('id')).order_by('type')
    type_counts = {item['type']: item['count'] for item in leads_by_type}
    
    # Лиды за последние 30 дней
    thirty_days_ago = timezone.now() - timedelta(days=30)
    recent_leads = queryset.filter(created_at__gte=thirty_days_ago).count()
    
    # Лиды за сегодня
    today = timezone.now().date()
    today_leads = queryset.filter(created_at__date=today).count()
    
    # График лидов по дням (последние 7 дней)
    daily_leads = []
    for i in range(6, -1, -1):
        date = timezone.now().date() - timedelta(days=i)
        count = queryset.filter(created_at__date=date).count()
        daily_leads.append({
            'date': date.strftime('%d.%m'),
            'count': count
        })
    
    # Топ городов
    top_cities = queryset.filter(city__isnull=False).values('city__name_ru').annotate(
        count=Count('id')
    ).order_by('-count')[:5]
    
    # Топ категорий
    top_categories = queryset.filter(category__isnull=False).values('category__name_ru').annotate(
        count=Count('id')
    ).order_by('-count')[:5]
    
    # Конверсия по статусам
    conversion_data = {
        'new': status_counts.get('NEW', 0),
        'confirmed': status_counts.get('CONFIRMED', 0),
        'paid': status_counts.get('PAID', 0),
        'done': status_counts.get('DONE', 0),
        'canceled': status_counts.get('CANCELED', 0),
    }
    
    # Среднее время обработки (от создания до оплаты)
    paid_leads = queryset.filter(status=Lead.LeadStatus.PAID, created_at__isnull=False)
    avg_processing_time = None
    if paid_leads.exists():
        # Упрощенный расчет - разница между created_at и updated_at для оплаченных
        avg_processing_time = paid_leads.aggregate(
            avg_time=Avg(
                F('updated_at') - F('created_at')
            )
        )['avg_time']
    
    # Последние лиды
    latest_leads = queryset.order_by('-created_at')[:10]
    
    context = {
        'total_leads': total_leads,
        'status_counts': status_counts,
        'type_counts': type_counts,
        'recent_leads': recent_leads,
        'today_leads': today_leads,
        'daily_leads': daily_leads,
        'top_cities': top_cities,
        'top_categories': top_categories,
        'conversion_data': conversion_data,
        'avg_processing_time': avg_processing_time,
        'latest_leads': latest_leads,
    }
    return render(request, 'srm_ui/dashboard.html', context)


@login_required
@require_http_methods(["POST"])
def bulk_action(request):
    """Массовые действия с лидами"""
    user = request.user
    action = request.POST.get("action")
    lead_ids_str = request.POST.get("lead_ids", "")
    
    if not lead_ids_str:
        messages.error(request, "Не выбрано ни одного лида")
        return redirect("srm_ui:lead_list")
    
    # Парсим ID из строки (разделены запятыми)
    lead_ids = [lid.strip() for lid in lead_ids_str.split(",") if lid.strip()]
    
    if not lead_ids:
        messages.error(request, "Не выбрано ни одного лида")
        return redirect("srm_ui:lead_list")
    
    # Получаем queryset с учетом прав доступа
    queryset = Lead.objects.filter(id__in=lead_ids)
    if hasattr(user, "role") and user.role == SrmUser.Roles.SCHOOL_MANAGER:
        if user.school:
            queryset = queryset.filter(school=user.school)
        else:
            queryset = Lead.objects.none()
    
    leads = list(queryset)
    
    if action == "change_status":
        new_status = request.POST.get("new_status")
        if new_status:
            count = 0
            for lead in leads:
                old_status = lead.status
                if old_status != new_status:
                    lead.status = new_status
                    lead.save()
                    LeadStatusHistory.objects.create(
                        lead=lead,
                        old_status=old_status,
                        new_status=new_status,
                        changed_by_user=user,
                        note=f"Массовое изменение статуса",
                    )
                    count += 1
            messages.success(request, f"Статус изменен для {count} лидов")
    
    elif action == "export_selected":
        # Экспорт выбранных лидов в CSV
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="selected_leads.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            "ID", "Тип", "Статус", "Имя", "Телефон", "ИИН", "WhatsApp", "Email",
            "Город", "Категория", "Автошкола", "Инструктор", "Дата создания"
        ])
        
        for lead in leads:
            writer.writerow([
                str(lead.id),
                lead.get_type_display(),
                lead.get_status_display(),
                lead.name,
                lead.phone,
                lead.iin or "",
                lead.whatsapp or "",
                lead.email or "",
                lead.city.name_ru if lead.city else "",
                lead.category.name_ru if lead.category else "",
                lead.school.name_ru if lead.school else "",
                lead.instructor.display_name if lead.instructor else "",
                lead.created_at.strftime("%d.%m.%Y %H:%M"),
            ])
        
        return response
    
    elif action == "add_payment_link":
        payment_link = request.POST.get("payment_link", "").strip()
        if payment_link:
            count = 0
            for lead in leads:
                lead.payment_link = payment_link
                lead.save()
                count += 1
            messages.success(request, f"Ссылка на оплату добавлена для {count} лидов")
    
    return redirect("srm_ui:lead_list")


@login_required
@require_http_methods(["GET"])
def lead_whatsapp_link(request, lead_id):
    """Получить WhatsApp ссылку для лида"""
    lead = get_object_or_404(Lead, id=lead_id)
    
    # Проверка прав доступа
    user = request.user
    if hasattr(user, "role") and user.role == SrmUser.Roles.SCHOOL_MANAGER:
        if not user.school or lead.school != user.school:
            return JsonResponse({"error": "Access denied"}, status=403)
    
    # Определяем номер WhatsApp
    whatsapp_phone = None
    if lead.type == Lead.LeadType.SCHOOL and lead.school:
        whatsapp_phone = lead.school.whatsapp_phone
    elif lead.type == Lead.LeadType.INSTRUCTOR and lead.instructor:
        # У инструктора может не быть WhatsApp, используем телефон клиента
        whatsapp_phone = lead.whatsapp or lead.phone
    elif lead.type == Lead.LeadType.TESTS:
        # Для тестов отправляем владельцу
        try:
            setting = ProjectSetting.objects.get(key="owner_whatsapp")
            whatsapp_phone = setting.value_json.get("phone", "")
        except ProjectSetting.DoesNotExist:
            pass
    
    if not whatsapp_phone:
        return JsonResponse({"error": "WhatsApp phone not found"}, status=404)
    
    # Генерируем сообщение
    message = _generate_whatsapp_message(lead)
    wa_link = _build_wa_link(whatsapp_phone, message)
    
    return JsonResponse({"wa_link": wa_link, "message": message})


@login_required
def notification_count(request):
    """API endpoint для получения количества непрочитанных уведомлений"""
    from .models import Notification
    count = Notification.objects.filter(user=request.user, is_read=False).count()
    return JsonResponse({"count": count})


@login_required
def notification_list(request):
    """Список уведомлений"""
    from .models import Notification
    notifications = Notification.objects.filter(user=request.user).order_by("-created_at")[:20]
    return render(request, "srm_ui/notifications.html", {"notifications": notifications})


@login_required
@require_http_methods(["POST"])
def mark_notification_read(request, notification_id):
    """Пометить уведомление как прочитанное"""
    from .models import Notification
    notification = get_object_or_404(Notification, id=notification_id, user=request.user)
    notification.is_read = True
    notification.save()
    return JsonResponse({"success": True})


@login_required
@require_http_methods(["POST"])
def mark_all_notifications_read(request):
    """Пометить все уведомления как прочитанные"""
    from .models import Notification
    Notification.objects.filter(user=request.user, is_read=False).update(is_read=True)
    return JsonResponse({"success": True})


@login_required
@require_http_methods(["POST"])
def add_comment(request, lead_id):
    """Добавить комментарий к лиду"""
    lead = get_object_or_404(Lead, id=lead_id)
    
    # Проверка прав доступа
    user = request.user
    if hasattr(user, "role") and user.role == SrmUser.Roles.SCHOOL_MANAGER:
        if not user.school or lead.school != user.school:
            return JsonResponse({"error": "Access denied"}, status=403)
    
    comment_text = request.POST.get("comment", "").strip()
    if not comment_text:
        return JsonResponse({"error": "Comment is required"}, status=400)
    
    from .models import LeadComment
    comment = LeadComment.objects.create(
        lead=lead,
        user=user,
        comment=comment_text
    )
    
    return JsonResponse({
        "success": True,
        "comment": {
            "id": comment.id,
            "comment": comment.comment,
            "user": comment.user.username,
            "created_at": comment.created_at.strftime("%d.%m.%Y %H:%M"),
        }
    })


@login_required
@require_http_methods(["GET"])
def get_comments(request, lead_id):
    """Получить комментарии к лиду (AJAX)"""
    lead = get_object_or_404(Lead, id=lead_id)
    
    # Проверка прав доступа
    user = request.user
    if hasattr(user, "role") and user.role == SrmUser.Roles.SCHOOL_MANAGER:
        if not user.school or lead.school != user.school:
            return JsonResponse({"error": "Access denied"}, status=403)
    
    from .models import LeadComment
    comments = LeadComment.objects.filter(lead=lead).order_by("-created_at")
    
    return JsonResponse({
        "comments": [
            {
                "id": c.id,
                "comment": c.comment,
                "user": c.user.username,
                "created_at": c.created_at.strftime("%d.%m.%Y %H:%M"),
            }
            for c in comments
        ]
    })


@login_required
def analytics(request):
    """Страница аналитики с конверсиями и графиками"""
    from analytics.models import AnalyticsEvent
    from django.db.models import Count, Q
    from datetime import timedelta
    
    user = request.user
    queryset = Lead.objects.all()
    
    # Фильтрация по правам доступа
    if hasattr(user, "role") and user.role == SrmUser.Roles.SCHOOL_MANAGER:
        if user.school:
            queryset = queryset.filter(school=user.school)
        else:
            queryset = Lead.objects.none()
    
    # Конверсия по этапам воронки
    total_bot_entries = AnalyticsEvent.objects.filter(event_name="bot_started").count()
    scenario_selections = AnalyticsEvent.objects.filter(event_name="flow_selected").count()
    card_views = AnalyticsEvent.objects.filter(
        Q(event_name="school_opened") | Q(event_name="instructor_opened")
    ).count()
    register_clicks = AnalyticsEvent.objects.filter(event_name="register_button_clicked").count()
    form_starts = AnalyticsEvent.objects.filter(event_name="lead_form_opened").count()
    form_submits = AnalyticsEvent.objects.filter(event_name="lead_submitted").count()
    whatsapp_opens = AnalyticsEvent.objects.filter(event_name="whatsapp_opened").count()
    paid_leads = queryset.filter(status=Lead.LeadStatus.PAID).count()
    completed_leads = queryset.filter(status=Lead.LeadStatus.DONE).count()
    
    # Расчет конверсий
    conversion_funnel = {
        "bot_entry": total_bot_entries,
        "scenario_selection": scenario_selections,
        "scenario_conversion": (scenario_selections / total_bot_entries * 100) if total_bot_entries > 0 else 0,
        "card_view": card_views,
        "card_conversion": (card_views / scenario_selections * 100) if scenario_selections > 0 else 0,
        "register_click": register_clicks,
        "register_conversion": (register_clicks / card_views * 100) if card_views > 0 else 0,
        "form_start": form_starts,
        "form_conversion": (form_starts / register_clicks * 100) if register_clicks > 0 else 0,
        "form_submit": form_submits,
        "submit_conversion": (form_submits / form_starts * 100) if form_starts > 0 else 0,
        "whatsapp_open": whatsapp_opens,
        "whatsapp_conversion": (whatsapp_opens / form_submits * 100) if form_submits > 0 else 0,
        "paid": paid_leads,
        "paid_conversion": (paid_leads / form_submits * 100) if form_submits > 0 else 0,
        "completed": completed_leads,
        "overall_conversion": (completed_leads / total_bot_entries * 100) if total_bot_entries > 0 else 0,
    }
    
    # График лидов по времени (последние 30 дней)
    daily_leads = []
    for i in range(29, -1, -1):
        date = timezone.now().date() - timedelta(days=i)
        count = queryset.filter(created_at__date=date).count()
        daily_leads.append({
            'date': date.strftime('%d.%m'),
            'count': count
        })
    
    # Статистика по источникам
    sources = queryset.values('source').annotate(count=Count('id')).order_by('-count')
    
    # Статистика по городам
    cities_stats = queryset.filter(city__isnull=False).values('city__name_ru').annotate(
        count=Count('id')
    ).order_by('-count')[:10]
    
    # Статистика по категориям
    categories_stats = queryset.filter(category__isnull=False).values('category__name_ru').annotate(
        count=Count('id')
    ).order_by('-count')[:10]
    
    import json
    
    context = {
        'conversion_funnel': conversion_funnel,
        'daily_leads': json.dumps(daily_leads),
        'sources': list(sources),
        'cities_stats': list(cities_stats),
        'categories_stats': list(categories_stats),
    }
    return render(request, 'srm_ui/analytics.html', context)


@login_required
def whatsapp_template_list(request):
    """Список WhatsApp шаблонов"""
    templates = WhatsAppTemplate.objects.all().order_by('scope', 'language')
    return render(request, 'srm_ui/whatsapp_templates.html', {'templates': templates})


@login_required
def whatsapp_template_edit(request, template_id=None):
    """Редактирование WhatsApp шаблона"""
    if template_id:
        template = get_object_or_404(WhatsAppTemplate, id=template_id)
    else:
        template = None
    
    if request.method == "POST":
        scope = request.POST.get("scope")
        language = request.POST.get("language")
        template_text = request.POST.get("template_text")
        is_active = request.POST.get("is_active") == "on"
        
        if template:
            template.scope = scope
            template.language = language
            template.template_text = template_text
            template.is_active = is_active
            template.save()
            messages.success(request, "Шаблон обновлен")
        else:
            template = WhatsAppTemplate.objects.create(
                scope=scope,
                language=language,
                template_text=template_text,
                is_active=is_active
            )
            messages.success(request, "Шаблон создан")
        
        return redirect("srm_ui:whatsapp_template_list")
    
    # Предпросмотр шаблона
    preview_data = None
    if template and request.GET.get("preview"):
        # Примерные данные для предпросмотра
        if template.scope == WhatsAppTemplate.Scope.SCHOOL_CLIENT_MESSAGE:
            preview_data = template.render(
                school_name="Smart Автошкола",
                intake_date="19 января",
                address="г. Атырау, Курмангазы 70Б",
                tariff_name="Базовый",
                tariff_price="90 000"
            )
        elif template.scope == WhatsAppTemplate.Scope.INSTRUCTOR_CLIENT_MESSAGE:
            preview_data = template.render(
                instructor_name="Конысов Бекали",
                gearbox="Автомат",
                tariff_name="1 час",
                tariff_price="8 000"
            )
        elif template.scope == WhatsAppTemplate.Scope.TESTS_OWNER_MESSAGE:
            preview_data = template.render(
                name="Иван Иванов",
                iin="123456789012",
                whatsapp="+77001234567",
                category="B",
                language=template.language
            )
    
    context = {
        'template': template,
        'scopes': WhatsAppTemplate.Scope.choices,
        'languages': [('RU', 'Русский'), ('KZ', 'Қазақша')],
        'preview_data': preview_data,
    }
    return render(request, 'srm_ui/whatsapp_template_edit.html', context)
